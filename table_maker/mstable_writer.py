from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import traceback

from .matched_arg import MatchedArg, PumpInfoGroup, PumpInfoArg
from .const_def import SalesPumpInfo

_ROW_ID = 8
_LIST_TABLE_SHEET = "设备一览表"

class MSTableWriter():
    def __init__(self) -> None:
        
        self._pos = [["B", "C", "D", None, None, None, None, None], 
            ["E", "F", "G", "H", None, "I", "J", 
            "K", None, "L", "M", "N", "O", "P", None,
            "Q", "R", "S", None, "T", "U"], 
            ["W", None, "Y", "Z", 
            None, None, "AC", None, "AE", "AF", "AG"], 
            ["AH", "AK", "AL", "AM"]
        ]

        self._pump_pos = [
            None, None, None, "X", "AN", None, None, "AC", None, None,
            None, None, "AO", None, None, None, None, None, None, "AV",
            "AW", "AX", "AY", "AZ", "BA", "BB", "BC", None, None, None,
            None, "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP",
            "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
            "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ",
            "CK", None, None
        ]

        self._formulars = [
            ("AA", lambda row: f"=MAX(R{row}, S{row})"), 
            ("AB", lambda row: f"=T{row}"),
            ("AD", lambda row: f"=AA{row}*AB{row}*I{row}/102/3600/AC{row}*100"),
            ("AP", lambda row: f"=VLOOKUP(Z{row},材质!B1:M15,2,0)"),
            ("AQ", lambda row: f"=VLOOKUP(Z{row},材质!D1:E15,2,0)"),
            ("AR", lambda row: f"=VLOOKUP(Z{row},材质!F1:G15,2,0)"),
            ("AS", lambda row: f"=VLOOKUP(Z{row},材质!H1:I15,2,0)"),
            ("AT", lambda row: f"=VLOOKUP(Z{row},材质!J1:K15,2,0)"),
            ("AU", lambda row: f"=VLOOKUP(Z{row},材质!L1:M15,2,0)"),
            ("BD", lambda row: f"=BC{row}*0.7"),
            ("BE", lambda row: f"=BC{row}*1.1"),
            ("BF", lambda row: f"=BC{row}*0.3"),
            ("BG", lambda row: f"=BC{row}*1.2"),
        ]

        self._from_func = {
            SalesPumpInfo.Flow.value: "AA",
            SalesPumpInfo.Lift.value: "AB",
            SalesPumpInfo.AxisEff.value: "AD",
            SalesPumpInfo.MatOfPump.value: "AP",
            SalesPumpInfo.MatOfWheel.value: "AQ",
            SalesPumpInfo.MatOfAxis.value: "AR",
            SalesPumpInfo.MatOfPumpRing.value: "AS",
            SalesPumpInfo.MatOfWheelRing.value: "AT",
            SalesPumpInfo.MatOfGuideVane.value: "AU",
            SalesPumpInfo.BestWorkRange0.value: "BD",
            SalesPumpInfo.BestWorkRange1.value: "BE",
            SalesPumpInfo.AllowWorkRange0.value: "BF",
            SalesPumpInfo.AllowWorkRange1.value: "BG",
        }
    
    def write_list(self, table_name: str, result: list[MatchedArg], pumps: list[PumpInfoGroup]):
        return self.write_table(table_name, result, pumps, "设备一览表")

    def set_pump_cell(self, index, pump: PumpInfoArg, ws: Worksheet, row_str: str):
        parg = pump.get_arg(index)
        ppos = self._pump_pos[index]
        if ppos is not None and parg.value is not None:
            if parg.unit is not None:
                ws[ppos + row_str].value = float(parg.value)
            else:
                ws[ppos + row_str].value = parg.value

    def read_table(self, table_name: str, cnt: int, pumps: list[PumpInfoArg], sheet="设备选型表 "):
        err_msg = None
        try:
            wb = load_workbook(table_name, read_only=True, data_only=True)
            ws = wb[sheet]

            for i in range(cnt):
                row_str = str(_ROW_ID + i)
                pump = PumpInfoArg()
                pumps.append(pump)

                for parg in pump.args:
                    aid = parg.index[1]
                    ppos = self._from_func.get(aid)
                    if ppos is None:
                        ppos = self._pump_pos[aid]
                    if ppos is None:
                        continue
                    cell_val = ws[ppos + row_str].value
                    if cell_val is not None:
                        parg.set_value(cell_val)
            wb.close()
            return None
        except Exception:
            traceback.print_exc() 
            if err_msg is None:
                err_msg = "写入错误，详见堆栈信息"
            return err_msg

    def write_table(self, table_name: str, result: list[MatchedArg], pumps: list[PumpInfoGroup] = None, sheet="设备选型表 "):
        err_msg = None
        try:
            wb = load_workbook(table_name, rich_text=True)
            ws = wb[sheet]
            cnt = len(result)
            ws.insert_rows(_ROW_ID, cnt)
            for i in range(cnt):
                row_str = str(_ROW_ID + i)
                args = result[i]
                pump = None
                if pumps is not None:
                    pump = pumps[i]

                for arg in args.found_args:
                    entry = args.get_arg(arg[0], arg[1])
                    pos = self._pos[arg[0]][arg[1]]
                    if pos is not None and entry.value is not None:
                        if entry.unit is not None:
                            ws[pos + row_str].value = float(entry.value)
                        else:
                            ws[pos + row_str].value = entry.value

                if pump is None:
                    continue

                pump_info = pump.pumps[pump.selected]
                if sheet == _LIST_TABLE_SHEET:
                    for col, form in self._formulars[0:3]:
                        ws[col + row_str] = form(_ROW_ID + i)
                    self.set_pump_cell(SalesPumpInfo.Prototype.value, pump_info, ws, row_str)
                    self.set_pump_cell(SalesPumpInfo.Efficience.value, pump_info, ws, row_str)
                    continue

                for col, form in self._formulars:
                    ws[col + row_str] = form(_ROW_ID + i)

                for parg in pump_info.args:
                    aid = parg.index[1]
                    ppos = self._pump_pos[aid]
                    if ppos is not None and parg.value is not None:
                        if parg.unit is not None:
                            ws[ppos + row_str].value = float(parg.value)
                        else:
                            ws[ppos + row_str].value = parg.value

            wb.save(table_name)
            wb.close()
            return err_msg
        
        except Exception:
            traceback.print_exc() 
            if err_msg is None:
                err_msg = "写入错误，详见堆栈信息"
            return err_msg